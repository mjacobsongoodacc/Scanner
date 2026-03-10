"""
NBA/NCAAB Arbitrage Scanner
Pulls odds from Kalshi (public API) + The Odds API (multi-book)
Compares cross-book and Kalshi-vs-book for arb opportunities and outputs to Excel.

Setup:
  pip install requests openpyxl
  Get a free API key at https://the-odds-api.com (500 req/month)

Usage:
  python arbitrage_scanner.py --odds-api-key YOUR_KEY [--sport nba|ncaab|both] [--stake 100]
"""

import argparse, json, math, re, requests, sys
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ─── API CONFIG ───────────────────────────────────────────────────────────────

KALSHI_BASE = "https://api.elections.kalshi.com/trade-api/v2"
KALSHI_API_KEY = "1abdab8d-2e42-4705-b12f-8a7be9e72d43"
ODDS_API_BASE = "https://api.the-odds-api.com/v4"

SPORT_MAP = {
    "nba": {
        "odds_api_key": "basketball_nba",
        "kalshi_series": {
            "spread": "KXNBASPREAD",
            "moneyline": "KXNBA",
        },
        "label": "NBA",
    },
    "ncaab": {
        "odds_api_key": "basketball_ncaab",
        "kalshi_series": {
            "spread": "KXNCAAMB1HSPREAD",
            "moneyline": "KXNCAAMB",
        },
        "label": "NCAAB",
    },
}

CITY_MAP = {
    "los angeles l": "lakers",
    "los angeles c": "clippers",
    "new york": "knicks",
    "brooklyn": "nets",
    "oklahoma city": "thunder",
    "golden state": "warriors",
    "san antonio": "spurs",
}

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def team_match(kalshi_team, odds_api_team):
    k = re.sub(r"[^a-z0-9 ]", "", kalshi_team.lower()).strip()
    o = re.sub(r"[^a-z0-9 ]", "", odds_api_team.lower()).strip()
    if o in k or k in o:
        return True
    k_words = k.split()
    o_words = o.split()
    if k_words[0] == o_words[0] and len(k_words[0]) > 3:
        return True
    mapped = CITY_MAP.get(k)
    if mapped:
        return mapped in o
    if any(w in o_words for w in k_words if len(w) > 4):
        return True
    return False


def parse_kalshi_spread_title(title):
    m = re.match(r"^(.+?)\s+wins?\s+by\s+over\s+([\d.]+)\s+points?\??$", title, re.IGNORECASE)
    if not m:
        return None, None
    return m.group(1).strip(), float(m.group(2))


def parse_kalshi_moneyline_title(title):
    if re.search(r"\b(?:by\s+over|spread|points?|total|over/under)\b", title, re.IGNORECASE):
        return None
    m = re.match(r"^(?:will\s+(?:the\s+)?)?(.+?)\s+(?:wins?|beat|defeat)(?:\s*\??\s*$)", title, re.IGNORECASE)
    if m:
        return re.sub(r"^the\s+", "", m.group(1), flags=re.IGNORECASE).strip()
    m = re.match(r"^(.+?)\s+to\s+win\s*\??\s*$", title, re.IGNORECASE)
    if m:
        return re.sub(r"^the\s+", "", m.group(1), flags=re.IGNORECASE).strip()
    m = re.match(r"^(.+?)\s+(?:vs?\.?)\s+(.+?)\s*\??\s*$", title, re.IGNORECASE)
    if m:
        return re.sub(r"^the\s+", "", m.group(1), flags=re.IGNORECASE).strip()
    return None


# ─── KALSHI FUNCTIONS ─────────────────────────────────────────────────────────

def fetch_kalshi_game_markets(sport="nba"):
    """Fetch game-specific spread and moneyline markets from Kalshi."""
    cfg = SPORT_MAP[sport]
    kalshi_series = cfg["kalshi_series"]
    all_markets = []

    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {KALSHI_API_KEY}",
    }

    for mkt_type, series_ticker in kalshi_series.items():
        cursor = ""
        pages = 0
        while pages < 5:
            params = {
                "status": "open",
                "with_nested_markets": "true",
                "limit": "200",
                "series_ticker": series_ticker,
            }
            if cursor:
                params["cursor"] = cursor

            try:
                resp = requests.get(f"{KALSHI_BASE}/events", params=params, headers=headers, timeout=15)
                resp.raise_for_status()
                data = resp.json()
            except Exception as e:
                print(f"[WARN] Kalshi fetch failed for {mkt_type} ({series_ticker}): {e}")
                break

            for ev in data.get("events", []):
                for mkt in ev.get("markets", []):
                    if mkt.get("status") in ("finalized", "closed"):
                        continue
                    yes_ask = mkt.get("yes_ask", 0)
                    no_ask = mkt.get("no_ask", 0)
                    yes_bid = mkt.get("yes_bid", 0)
                    no_bid = mkt.get("no_bid", 0)
                    if not yes_ask and not no_ask:
                        continue

                    vol = mkt.get("volume", 0)
                    yes_ba = (yes_ask - yes_bid) if yes_ask and yes_bid else 99
                    no_ba = (no_ask - no_bid) if no_ask and no_bid else 99

                    if vol == 0:
                        continue
                    if yes_ba > 15 and no_ba > 15:
                        continue

                    team = None
                    spread = None
                    effective_type = mkt_type
                    spread_team, spread_val = parse_kalshi_spread_title(mkt.get("title", ""))
                    if spread_team is not None:
                        effective_type = "spread"
                        team = spread_team
                        spread = spread_val
                    elif mkt_type == "moneyline":
                        team = parse_kalshi_moneyline_title(mkt.get("title", ""))

                    all_markets.append({
                        "event_ticker": ev.get("event_ticker"),
                        "event_title": ev.get("title"),
                        "market_ticker": mkt.get("ticker"),
                        "market_title": mkt.get("title"),
                        "type": effective_type,
                        "yes_bid": yes_bid,
                        "yes_ask": yes_ask,
                        "no_bid": no_bid,
                        "no_ask": no_ask,
                        "volume": vol,
                        "yes_ba_spread": yes_ba,
                        "no_ba_spread": no_ba,
                        "close_time": mkt.get("close_time"),
                        "parsed_team": team,
                        "parsed_spread": spread,
                    })

            cursor = data.get("cursor", "")
            pages += 1
            if not cursor:
                break

    return all_markets


def kalshi_taker_fee(cents, contracts=1):
    """Total taker fee rounded up to cents for a whole-contract order."""
    if not cents or cents <= 0 or cents >= 100 or not contracts or contracts <= 0:
        return 0
    p = cents / 100
    return math.ceil(0.07 * contracts * p * (1 - p) * 100) / 100


def kalshi_fee_per_contract(cents):
    """One-contract taker fee, useful for UI-style inspection."""
    return kalshi_taker_fee(cents, 1)


def kalshi_cost_for_contracts(cents, contracts):
    if not contracts or contracts <= 0:
        return 0
    return contracts * (cents / 100) + kalshi_taker_fee(cents, contracts)


def kalshi_cents_to_decimal(cents):
    """Convert Kalshi cents to decimal odds, accounting for taker fee."""
    if not cents or cents <= 0 or cents >= 100:
        return None
    total_cost = kalshi_cost_for_contracts(cents, 1)
    if total_cost >= 1:
        return None
    return 1 / total_cost


def kalshi_cents_to_decimal_raw(cents):
    """Convert Kalshi cents to decimal odds WITHOUT fee adjustment."""
    if not cents or cents <= 0 or cents >= 100:
        return None
    return 100 / cents


def kalshi_cents_to_american(cents):
    """Convert Kalshi cents to American odds, accounting for taker fee."""
    dec = kalshi_cents_to_decimal(cents)
    if not dec:
        return None
    if dec >= 2.0:
        return round((dec - 1) * 100)
    else:
        return round(-100 / (dec - 1))


def is_whole_number(value, tol=1e-9):
    return abs(value - round(value)) < tol


def size_kalshi_vs_book(kalshi_cents, book_dec, stake):
    if not kalshi_cents or not book_dec or not stake or stake <= 0:
        return None

    def total_outlay(contracts):
        kalshi_cost = kalshi_cost_for_contracts(kalshi_cents, contracts)
        book_bet = contracts / book_dec
        return kalshi_cost + book_bet

    lo, hi = 0, 1
    while total_outlay(hi) <= stake:
        hi *= 2

    while lo + 1 < hi:
        mid = (lo + hi) // 2
        if total_outlay(mid) <= stake:
            lo = mid
        else:
            hi = mid

    contracts = lo
    if contracts < 1:
        return None

    fee = kalshi_taker_fee(kalshi_cents, contracts)
    kalshi_cost = kalshi_cost_for_contracts(kalshi_cents, contracts)
    book_bet = contracts / book_dec
    used_stake = kalshi_cost + book_bet
    profit = contracts - used_stake

    return {
        "contracts": contracts,
        "fee": fee,
        "kalshi_cost": kalshi_cost,
        "book_bet": book_bet,
        "used_stake": used_stake,
        "unused_stake": max(0, stake - used_stake),
        "profit": profit,
        "payout": contracts,
        "kalshi_decimal": contracts / kalshi_cost,
    }


def build_stake_plan(dec_a, dec_b, stake, kalshi_leg=None):
    if not dec_a or not dec_b or not stake or stake <= 0:
        return None

    if not kalshi_leg:
        imp_sum = 1 / dec_a + 1 / dec_b
        bet_a = (1 / dec_a / imp_sum) * stake
        bet_b = (1 / dec_b / imp_sum) * stake
        payout_a = bet_a * dec_a
        payout_b = bet_b * dec_b
        profit = min(payout_a, payout_b) - stake
        return {
            "imp_sum": imp_sum,
            "roi": (1 - imp_sum) * 100,
            "bet_a": bet_a,
            "bet_b": bet_b,
            "payout_a": payout_a,
            "payout_b": payout_b,
            "profit": profit,
            "used_stake": stake,
            "unused_stake": 0,
            "decimal_a": dec_a,
            "decimal_b": dec_b,
            "american_a": None,
            "american_b": None,
            "kalshi_contracts_a": None,
            "kalshi_contracts_b": None,
            "kalshi_fee_a": None,
            "kalshi_fee_b": None,
        }

    sized = size_kalshi_vs_book(
        kalshi_leg["cents"],
        dec_b if kalshi_leg["position"] == "a" else dec_a,
        stake,
    )
    if not sized:
        return None

    imp_sum = sized["used_stake"] / sized["payout"]
    common = {
        "imp_sum": imp_sum,
        "roi": (1 - imp_sum) * 100,
        "profit": sized["profit"],
        "used_stake": sized["used_stake"],
        "unused_stake": sized["unused_stake"],
    }

    if kalshi_leg["position"] == "a":
        return {
            **common,
            "bet_a": sized["kalshi_cost"],
            "bet_b": sized["book_bet"],
            "payout_a": sized["payout"],
            "payout_b": sized["payout"],
            "decimal_a": sized["kalshi_decimal"],
            "decimal_b": dec_b,
            "american_a": decimal_to_american(sized["kalshi_decimal"]),
            "american_b": None,
            "kalshi_contracts_a": sized["contracts"],
            "kalshi_contracts_b": None,
            "kalshi_fee_a": sized["fee"],
            "kalshi_fee_b": None,
        }

    return {
        **common,
        "bet_a": sized["book_bet"],
        "bet_b": sized["kalshi_cost"],
        "payout_a": sized["payout"],
        "payout_b": sized["payout"],
        "decimal_a": dec_a,
        "decimal_b": sized["kalshi_decimal"],
        "american_a": None,
        "american_b": decimal_to_american(sized["kalshi_decimal"]),
        "kalshi_contracts_a": None,
        "kalshi_contracts_b": sized["contracts"],
        "kalshi_fee_a": None,
        "kalshi_fee_b": sized["fee"],
    }


# ─── THE ODDS API FUNCTIONS ──────────────────────────────────────────────────

def fetch_odds_api_games(api_key, sport="nba"):
    """Fetch moneyline + spread odds from The Odds API."""
    cfg = SPORT_MAP[sport]
    url = f"{ODDS_API_BASE}/sports/{cfg['odds_api_key']}/odds"
    params = {
        "apiKey": api_key,
        "regions": "us",
        "markets": "h2h,spreads",
        "oddsFormat": "american",
    }

    try:
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        remaining = resp.headers.get("x-requests-remaining", "?")
        print(f"[INFO] Odds API requests remaining: {remaining}")
        return resp.json()
    except Exception as e:
        print(f"[WARN] Odds API fetch failed: {e}")
        return []


def parse_odds_api_games(games):
    parsed = []
    for game in games:
        home = game.get("home_team", "")
        away = game.get("away_team", "")
        start = game.get("commence_time", "")
        bookmakers = game.get("bookmakers", [])

        book_odds = {}
        spread_odds = {}
        for bk in bookmakers:
            book_name = bk.get("title", bk.get("key", "Unknown"))
            for mkt in bk.get("markets", []):
                if mkt.get("key") == "h2h":
                    outcomes = mkt.get("outcomes", [])
                    home_odds = None
                    away_odds = None
                    for oc in outcomes:
                        if oc["name"] == home:
                            home_odds = oc["price"]
                        elif oc["name"] == away:
                            away_odds = oc["price"]
                    if home_odds and away_odds:
                        book_odds[book_name] = {"home": home_odds, "away": away_odds}
                elif mkt.get("key") == "spreads":
                    lines = []
                    for oc in mkt.get("outcomes", []):
                        lines.append({
                            "name": oc["name"],
                            "point": oc.get("point", 0),
                            "price": oc["price"],
                        })
                    if lines:
                        spread_odds[book_name] = lines

        parsed.append({
            "home": home,
            "away": away,
            "commence_time": start,
            "book_odds": book_odds,
            "spread_odds": spread_odds,
        })
    return parsed


# ─── ARBITRAGE MATH ──────────────────────────────────────────────────────────

def american_to_decimal(american):
    if american > 0:
        return american / 100 + 1
    elif american < 0:
        return 100 / abs(american) + 1
    return None


def decimal_to_american(decimal):
    if not decimal or decimal <= 1:
        return None
    if decimal >= 2:
        return round((decimal - 1) * 100)
    return round(-100 / (decimal - 1))


def find_arbs(odds_games, kalshi_markets, stake=100, near_arb_threshold=1.03):
    opportunities = []
    best_imp_sum = float("inf")
    best_imp_detail = None

    def _record(imp_sum, dec_a, dec_b, side_a, book_a, am_a, side_b, book_b, am_b, game,
                km_ticker="", mkt_type="h2h", km_volume=None, km_ba_spread=None, kalshi_leg=None):
        nonlocal best_imp_sum, best_imp_detail
        stake_plan = build_stake_plan(dec_a, dec_b, stake, kalshi_leg)
        if not stake_plan:
            return

        exact_imp_sum = stake_plan["imp_sum"]
        if exact_imp_sum < best_imp_sum:
            best_imp_sum = exact_imp_sum
            best_imp_detail = f"{game['away']} @ {game['home']}: {side_a} ({book_a}) vs {side_b} ({book_b})"
        if exact_imp_sum < near_arb_threshold:
            roi = stake_plan["roi"]
            bet_a = stake_plan["bet_a"]
            bet_b = stake_plan["bet_b"]
            profit = stake_plan["profit"]

            confidence = "high"
            if km_ticker:
                vol = km_volume or 0
                ba = km_ba_spread if km_ba_spread is not None else 99
                if vol < 100 or ba > 5:
                    confidence = "low"
                elif vol < 500 or ba > 3:
                    confidence = "medium"
            if roi > 10:
                confidence = "low"
            elif roi > 5 and confidence != "low":
                confidence = "medium"

            opportunities.append({
                "game": f"{game['away']} @ {game['home']}",
                "commence": game["commence_time"],
                "side_a": side_a, "side_a_book": book_a,
                "side_a_odds_dec": round(stake_plan["decimal_a"], 4), "side_a_american": stake_plan["american_a"] if stake_plan["american_a"] is not None else am_a,
                "side_b": side_b, "side_b_book": book_b,
                "side_b_odds_dec": round(stake_plan["decimal_b"], 4), "side_b_american": stake_plan["american_b"] if stake_plan["american_b"] is not None else am_b,
                "implied_sum": round(exact_imp_sum, 6),
                "roi_pct": round(roi, 3),
                "bet_a": round(bet_a, 2), "bet_b": round(bet_b, 2),
                "used_stake": round(stake_plan["used_stake"], 2),
                "unused_stake": round(stake_plan["unused_stake"], 2),
                "guaranteed_profit": round(profit, 2),
                "is_true_arb": exact_imp_sum < 1.0 and profit > 0,
                "kalshi_ticker": km_ticker,
                "market_type": mkt_type,
                "confidence": confidence,
                "kalshi_volume": km_volume,
                "kalshi_ba_spread": km_ba_spread,
                "kalshi_contracts_a": stake_plan["kalshi_contracts_a"],
                "kalshi_contracts_b": stake_plan["kalshi_contracts_b"],
                "kalshi_fee_a": stake_plan["kalshi_fee_a"],
                "kalshi_fee_b": stake_plan["kalshi_fee_b"],
            })

    for game in odds_games:
        home = game["home"]
        away = game["away"]

        # ── H2H: Best-line across ALL books ──
        best_home_dec, best_home_book, best_home_am = 0, "", 0
        best_away_dec, best_away_book, best_away_am = 0, "", 0
        for bn, odds in game["book_odds"].items():
            hd = american_to_decimal(odds["home"])
            ad = american_to_decimal(odds["away"])
            if hd and hd > best_home_dec:
                best_home_dec, best_home_book, best_home_am = hd, bn, odds["home"]
            if ad and ad > best_away_dec:
                best_away_dec, best_away_book, best_away_am = ad, bn, odds["away"]
        if best_home_dec and best_away_dec and best_home_book != best_away_book:
            imp_sum = 1 / best_home_dec + 1 / best_away_dec
            _record(imp_sum, best_home_dec, best_away_dec,
                    f"{home} ML", best_home_book, best_home_am,
                    f"{away} ML", best_away_book, best_away_am, game)

        # ── H2H: Pairwise book-vs-book ──
        book_names = list(game["book_odds"].keys())
        for i in range(len(book_names)):
            for j in range(i + 1, len(book_names)):
                b1, b2 = book_names[i], book_names[j]
                o1, o2 = game["book_odds"][b1], game["book_odds"][b2]
                d1_home = american_to_decimal(o1["home"])
                d1_away = american_to_decimal(o1["away"])
                d2_home = american_to_decimal(o2["home"])
                d2_away = american_to_decimal(o2["away"])
                if d1_home and d2_away:
                    imp_sum = 1/d1_home + 1/d2_away
                    _record(imp_sum, d1_home, d2_away,
                            f"{home} ML", b1, o1["home"],
                            f"{away} ML", b2, o2["away"], game)
                if d2_home and d1_away:
                    imp_sum = 1/d2_home + 1/d1_away
                    _record(imp_sum, d2_home, d1_away,
                            f"{home} ML", b2, o2["home"],
                            f"{away} ML", b1, o1["away"], game)

        # ── MONEYLINE: Kalshi vs Books ──
        matched_kalshi_ml = [km for km in kalshi_markets
                             if km["type"] == "moneyline" and km["parsed_team"]
                             and (team_match(km["parsed_team"], home) or team_match(km["parsed_team"], away))]

        for km in matched_kalshi_ml:
            kalshi_is_home = team_match(km["parsed_team"], home)
            kalshi_team = home if kalshi_is_home else away
            opp_team = away if kalshi_is_home else home

            yes_dec = kalshi_cents_to_decimal(km["yes_ask"])
            no_dec = kalshi_cents_to_decimal(km["no_ask"])
            yes_am = kalshi_cents_to_american(km["yes_ask"])
            no_am = kalshi_cents_to_american(km["no_ask"])

            # Kalshi YES (team wins) vs best sportsbook ML on opponent
            best_opp_ml_dec, best_opp_ml_book, best_opp_ml_am = 0, "", 0
            for bn, odds in game["book_odds"].items():
                opp_am = odds["away"] if kalshi_is_home else odds["home"]
                opp_dec = american_to_decimal(opp_am)
                if opp_dec and opp_dec > best_opp_ml_dec:
                    best_opp_ml_dec, best_opp_ml_book, best_opp_ml_am = opp_dec, bn, opp_am
            if yes_dec and best_opp_ml_dec:
                imp_sum = 1/yes_dec + 1/best_opp_ml_dec
                _record(imp_sum, yes_dec, best_opp_ml_dec,
                        f"{kalshi_team} ML (Kalshi YES)", "Kalshi", yes_am,
                        f"{opp_team} ML", best_opp_ml_book, best_opp_ml_am,
                        game, km["market_ticker"], "h2h",
                        km["volume"], km.get("yes_ba_spread"),
                        {"position": "a", "cents": km["yes_ask"]})

            # Sportsbook ML on team vs Kalshi NO (team loses)
            best_same_ml_dec, best_same_ml_book, best_same_ml_am = 0, "", 0
            for bn, odds in game["book_odds"].items():
                same_am = odds["home"] if kalshi_is_home else odds["away"]
                same_dec = american_to_decimal(same_am)
                if same_dec and same_dec > best_same_ml_dec:
                    best_same_ml_dec, best_same_ml_book, best_same_ml_am = same_dec, bn, same_am
            if best_same_ml_dec and no_dec:
                imp_sum = 1/best_same_ml_dec + 1/no_dec
                _record(imp_sum, best_same_ml_dec, no_dec,
                        f"{kalshi_team} ML", best_same_ml_book, best_same_ml_am,
                        f"{opp_team} ML (Kalshi NO)", "Kalshi", no_am,
                        game, km["market_ticker"], "h2h",
                        km["volume"], km.get("no_ba_spread"),
                        {"position": "b", "cents": km["no_ask"]})

            # Pairwise: Kalshi vs each individual book
            for bn, odds in game["book_odds"].items():
                opp_am = odds["away"] if kalshi_is_home else odds["home"]
                opp_dec = american_to_decimal(opp_am)
                same_am = odds["home"] if kalshi_is_home else odds["away"]
                same_dec = american_to_decimal(same_am)

                if yes_dec and opp_dec:
                    imp_sum = 1/yes_dec + 1/opp_dec
                    _record(imp_sum, yes_dec, opp_dec,
                            f"{kalshi_team} ML (Kalshi YES)", "Kalshi", yes_am,
                            f"{opp_team} ML", bn, opp_am,
                            game, km["market_ticker"], "h2h",
                            km["volume"], km.get("yes_ba_spread"),
                            {"position": "a", "cents": km["yes_ask"]})
                if same_dec and no_dec:
                    imp_sum = 1/same_dec + 1/no_dec
                    _record(imp_sum, same_dec, no_dec,
                            f"{kalshi_team} ML", bn, same_am,
                            f"{opp_team} ML (Kalshi NO)", "Kalshi", no_am,
                            game, km["market_ticker"], "h2h",
                            km["volume"], km.get("no_ba_spread"),
                            {"position": "b", "cents": km["no_ask"]})

        # ── SPREAD: Kalshi vs Books ──
        spread_books = game.get("spread_odds", {})
        matched_kalshi = [km for km in kalshi_markets
                          if km["type"] == "spread" and km["parsed_team"] and km["parsed_spread"] is not None
                          and (team_match(km["parsed_team"], home) or team_match(km["parsed_team"], away))]

        for km in matched_kalshi:
            kalshi_is_home = team_match(km["parsed_team"], home)
            kalshi_team = home if kalshi_is_home else away
            opp_team = away if kalshi_is_home else home
            spread_val = km["parsed_spread"]

            yes_dec = kalshi_cents_to_decimal(km["yes_ask"])
            no_dec = kalshi_cents_to_decimal(km["no_ask"])
            yes_am = kalshi_cents_to_american(km["yes_ask"])
            no_am = kalshi_cents_to_american(km["no_ask"])

            for book_name, lines in spread_books.items():
                for line in lines:
                    if abs(abs(line["point"]) - spread_val) > 0.01:
                        continue
                    book_dec = american_to_decimal(line["price"])
                    if not book_dec:
                        continue

                    is_opposite = line["name"] == opp_team
                    is_same = line["name"] == kalshi_team

                    if is_opposite and yes_dec and not is_whole_number(spread_val):
                        imp_sum = 1/yes_dec + 1/book_dec
                        _record(imp_sum, yes_dec, book_dec,
                                f"{kalshi_team} -{spread_val} (Kalshi YES)", "Kalshi", yes_am,
                                f"{opp_team} +{spread_val}", book_name, line["price"],
                                game, km["market_ticker"], "spread",
                                km["volume"], km.get("yes_ba_spread"),
                                {"position": "a", "cents": km["yes_ask"]})

                    if is_same and no_dec:
                        imp_sum = 1/book_dec + 1/no_dec
                        _record(imp_sum, book_dec, no_dec,
                                f"{kalshi_team} -{spread_val}", book_name, line["price"],
                                f"{opp_team} +{spread_val} (Kalshi NO)", "Kalshi", no_am,
                                game, km["market_ticker"], "spread",
                                km["volume"], km.get("no_ba_spread"),
                                {"position": "b", "cents": km["no_ask"]})

            # ── SPREAD: Best-line across books for this spread value ──
            best_cover_dec, best_cover_book, best_cover_am = 0, "", 0
            best_opp_dec, best_opp_book, best_opp_am = 0, "", 0
            for book_name, lines in spread_books.items():
                for line in lines:
                    if abs(abs(line["point"]) - spread_val) > 0.01:
                        continue
                    dec = american_to_decimal(line["price"])
                    if not dec:
                        continue
                    is_same = line["name"] == kalshi_team
                    if is_same and dec > best_cover_dec:
                        best_cover_dec, best_cover_book, best_cover_am = dec, book_name, line["price"]
                    if (not is_same) and line["name"] == opp_team and dec > best_opp_dec:
                        best_opp_dec, best_opp_book, best_opp_am = dec, book_name, line["price"]

            if best_opp_dec and yes_dec and not is_whole_number(spread_val):
                imp_sum = 1/yes_dec + 1/best_opp_dec
                _record(imp_sum, yes_dec, best_opp_dec,
                        f"{kalshi_team} -{spread_val} (Kalshi YES)", "Kalshi", yes_am,
                        f"{opp_team} +{spread_val}", best_opp_book, best_opp_am,
                        game, km["market_ticker"], "spread",
                        km["volume"], km.get("yes_ba_spread"),
                        {"position": "a", "cents": km["yes_ask"]})

            if best_cover_dec and no_dec:
                imp_sum = 1/best_cover_dec + 1/no_dec
                _record(imp_sum, best_cover_dec, no_dec,
                        f"{kalshi_team} -{spread_val}", best_cover_book, best_cover_am,
                        f"{opp_team} +{spread_val} (Kalshi NO)", "Kalshi", no_am,
                        game, km["market_ticker"], "spread",
                        km["volume"], km.get("no_ba_spread"),
                        {"position": "b", "cents": km["no_ask"]})

        # ── SPREAD: Book-vs-book ──
        spread_by_line = {}
        for book_name, lines in spread_books.items():
            for line in lines:
                key = f"{abs(line['point']):.1f}"
                spread_by_line.setdefault(key, []).append({**line, "book": book_name})
        for entries in spread_by_line.values():
            for i in range(len(entries)):
                for j in range(i + 1, len(entries)):
                    a, b = entries[i], entries[j]
                    if a["book"] == b["book"] or a["name"] == b["name"]:
                        continue
                    dec_a = american_to_decimal(a["price"])
                    dec_b = american_to_decimal(b["price"])
                    if not dec_a or not dec_b:
                        continue
                    spread_val = abs(a["point"])
                    if is_whole_number(spread_val):
                        continue
                    imp_sum = 1/dec_a + 1/dec_b
                    _record(imp_sum, dec_a, dec_b,
                            f"{a['name']} {a['point']:+g}", a["book"], a["price"],
                            f"{b['name']} {b['point']:+g}", b["book"], b["price"],
                            game, "", "spread")

    opportunities.sort(key=lambda x: x["implied_sum"])
    return opportunities, best_imp_sum, best_imp_detail


# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────

def write_excel(opportunities, all_games, stake, filename="arb_scan_results.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Arb Opportunities"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1B5E20")
    arb_fill = PatternFill("solid", fgColor="C8E6C9")
    near_fill = PatternFill("solid", fgColor="FFF3E0")
    black_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws["A1"] = "Arbitrage Scanner Results"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Stake: ${stake}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True)

    headers = [
        "Game", "Bet Type", "Tip-Off", "Side A", "Book A", "American A", "Decimal A", "Implied Prob A",
        "Side B", "Book B", "American B", "Decimal B", "Implied Prob B",
        "Implied Sum", "ROI %", "Bet A ($)", "Bet B ($)", "Used Stake ($)", "Cash Left ($)",
        "Guaranteed Profit", "True Arb?", "Confidence", "Kalshi Ticker",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for i, opp in enumerate(opportunities):
        row = 5 + i
        bet_type = "Moneyline" if opp["market_type"] == "h2h" else "Spread" if opp["market_type"] == "spread" else opp["market_type"]
        imp_prob_a = round(100 / opp["side_a_odds_dec"], 1) if opp["side_a_odds_dec"] else ""
        imp_prob_b = round(100 / opp["side_b_odds_dec"], 1) if opp["side_b_odds_dec"] else ""
        vals = [
            opp["game"], bet_type,
            opp["commence"][:16].replace("T", " ") if opp["commence"] else "",
            opp["side_a"], opp["side_a_book"], opp["side_a_american"],
            opp["side_a_odds_dec"], f"{imp_prob_a}%",
            opp["side_b"], opp["side_b_book"], opp["side_b_american"],
            opp["side_b_odds_dec"], f"{imp_prob_b}%",
            opp["implied_sum"], opp["roi_pct"],
            opp["bet_a"], opp["bet_b"],
            opp["used_stake"], opp["unused_stake"],
            opp["guaranteed_profit"],
            "YES" if opp["is_true_arb"] else "near",
            opp["confidence"],
            opp["kalshi_ticker"],
        ]
        fill = arb_fill if opp["is_true_arb"] else near_fill
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = black_font
            cell.border = thin_border
            cell.fill = fill

    widths = [28, 12, 18, 24, 14, 12, 10, 12, 24, 14, 12, 10, 12, 12, 8, 12, 12, 12, 12, 12, 8, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if not opportunities:
        ws.cell(row=5, column=1, value="No arbitrage opportunities found at this time.")
        ws.cell(row=5, column=1).font = Font(name="Arial", italic=True, size=11)

    wb.save(filename)
    print(f"\n[OK] Results saved to {filename}")
    return filename


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="NBA/NCAAB Arbitrage Scanner")
    parser.add_argument("--odds-api-key", required=True, help="The Odds API key")
    parser.add_argument("--sport", default="nba", choices=["nba", "ncaab", "both"])
    parser.add_argument("--stake", type=float, default=100)
    parser.add_argument("--output", default="arb_scan_results.xlsx")
    args = parser.parse_args()

    sports = ["nba", "ncaab"] if args.sport == "both" else [args.sport]
    all_opps = []
    all_games = []

    for sport in sports:
        label = SPORT_MAP[sport]["label"]
        print(f"\n{'='*60}")
        print(f"  Scanning {label}")
        print(f"{'='*60}")

        print(f"[1/3] Fetching Kalshi {label} game markets...")
        kalshi = fetch_kalshi_game_markets(sport)
        kalshi_spread = [k for k in kalshi if k["type"] == "spread"]
        kalshi_ml = [k for k in kalshi if k["type"] == "moneyline"]
        print(f"       Found {len(kalshi)} Kalshi markets ({len(kalshi_spread)} spread, {len(kalshi_ml)} moneyline)")

        print(f"[2/3] Fetching {label} odds from The Odds API...")
        raw_games = fetch_odds_api_games(args.odds_api_key, sport)
        games = parse_odds_api_games(raw_games)
        print(f"       Found {len(games)} games across {sum(len(g['book_odds']) for g in games)} book lines")
        all_games.extend(games)

        print(f"[3/3] Scanning for arbitrage opportunities...")
        opps, best_imp, best_detail = find_arbs(games, kalshi, args.stake)
        true_arbs = [o for o in opps if o["is_true_arb"]]
        near_arbs = [o for o in opps if not o["is_true_arb"]]
        kalshi_arbs = [o for o in opps if o["kalshi_ticker"]]
        print(f"       Found {len(true_arbs)} true arbs, {len(near_arbs)} near-arbs")
        print(f"       Kalshi cross-exchange: {len(kalshi_arbs)}")
        if best_detail:
            gap = (best_imp - 1) * 100
            status = "ARB!" if best_imp < 1 else f"{gap:.2f}% away"
            print(f"       Best: {best_detail} (imp={best_imp:.6f}, {status})")
        all_opps.extend(opps)

    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    print(f"  Total games scanned:  {len(all_games)}")
    true_arbs_all = [o for o in all_opps if o["is_true_arb"]]
    near_arbs_all = [o for o in all_opps if not o["is_true_arb"]]
    kalshi_arbs_all = [o for o in all_opps if o["kalshi_ticker"]]
    print(f"  True arbs:            {len(true_arbs_all)}")
    print(f"  Near-arbs (<3%):      {len(near_arbs_all)}")
    print(f"  Kalshi cross-exchange: {len(kalshi_arbs_all)}")
    ml_count = len([o for o in all_opps if o["market_type"] == "h2h"])
    spread_count = len([o for o in all_opps if o["market_type"] == "spread"])
    print(f"  Moneyline opps:       {ml_count}")
    print(f"  Spread opps:          {spread_count}")

    if true_arbs_all:
        best = true_arbs_all[0]
        bet_type = "Moneyline" if best["market_type"] == "h2h" else "Spread"
        print(f"\n  BEST TRUE ARB [{bet_type}]:")
        print(f"    {best['game']}")
        print(f"    {best['side_a']} @ {best['side_a_book']} ({best['side_a_american']})")
        print(f"    {best['side_b']} @ {best['side_b_book']} ({best['side_b_american']})")
        print(f"    ROI: {best['roi_pct']:.3f}%  |  Profit: ${best['guaranteed_profit']:.2f} on ${args.stake}")
        print(f"    Confidence: {best['confidence']}")
    elif near_arbs_all:
        best = near_arbs_all[0]
        gap = (best["implied_sum"] - 1) * 100
        bet_type = "Moneyline" if best["market_type"] == "h2h" else "Spread"
        print(f"\n  CLOSEST TO ARB [{bet_type}]:")
        print(f"    {best['game']}")
        print(f"    {best['side_a']} @ {best['side_a_book']} ({best['side_a_american']})")
        print(f"    {best['side_b']} @ {best['side_b_book']} ({best['side_b_american']})")
        print(f"    Implied sum: {best['implied_sum']:.6f}  |  Gap: {gap:.2f}%")
    else:
        print("\n  No opportunities found right now.")

    write_excel(all_opps, all_games, args.stake, args.output)


if __name__ == "__main__":
    main()
